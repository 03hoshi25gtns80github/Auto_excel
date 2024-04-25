import openpyxl
import os
from tkinter import filedialog
from tkinter import Tk

def update_excel_file(input_file_path):
    # Load the workbook
    wb = openpyxl.load_workbook(input_file_path)
    
    # Get the sheet names
    sheet_names = wb.sheetnames
    
    # Get the content of Sheet1
    sheet1 = wb[sheet_names[0]]
    
    def copy_non_empty_cells(sheet1, sheet):
        for row in sheet.iter_rows(max_row=9):  # Only iterate through rows 1-9
            for cell in row:
                sheet1_cell_value = sheet1[cell.coordinate].value
                # If the cell in the other sheet is empty and the corresponding cell in sheet1 is not empty
                if not cell.value and sheet1_cell_value:
                    cell.value = sheet1_cell_value
        
        # Copy cell J1 from sheet1 to the other sheet
        sheet['J1'].value = sheet1['J1'].value
    
    # Iterate through the other sheets and copy non-empty cells from sheet1 with updated requirements
    for sheet_name in sheet_names[1:]:  # Skip Sheet1
        sheet = wb[sheet_name]
        copy_non_empty_cells(sheet1, sheet)
    
    # Create the output file path
    directory, input_file_name = os.path.split(input_file_path)
    output_file_name = f'updated_{input_file_name}'
    output_file_path = os.path.join(directory, output_file_name)
    
    # Save the workbook to the output file
    wb.save(output_file_path)

def main():
    # Hide the root Tkinter window
    root = Tk()
    root.withdraw()
    
    # Prompt the user to select the input Excel file
    input_file_path = filedialog.askopenfilename(title='Select Excel File', filetypes=[('Excel Files', '*.xlsx')])
    
    # Check if a file was selected
    if input_file_path:
        update_excel_file(input_file_path)
    else:
        print('No file selected.')

if __name__ == '__main__':
    main()
